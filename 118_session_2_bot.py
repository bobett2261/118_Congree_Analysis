import requests
from lxml import etree
import openpyxl

def fetch_roll_call_data(roll_call_number, year=2024):
    """Fetch and parse roll call vote data from the XML URL."""
    url = f"https://clerk.house.gov/evs/{year}/roll{str(roll_call_number).zfill(3)}.xml"
    response = requests.get(url)
    if response.status_code == 200:
        return etree.XML(response.content)
    else:
        print(f"Failed to fetch data for roll call {roll_call_number}")
        return None

def parse_xml_data(xml_data):
    """Extract relevant data from the XML."""
    metadata = xml_data.find(".//vote-metadata")
    totals = xml_data.find(".//vote-totals/totals-by-vote")
    data = {
        'Majority': metadata.findtext("majority", default='N/A'),
        'Congress': metadata.findtext("congress", default='N/A'),
        'Session': metadata.findtext("session", default='N/A'),
        'Chamber': metadata.findtext("chamber", default='N/A'),
        'Roll Call Number': metadata.findtext("rollcall-num", default='N/A'),
        'Legislation Number': metadata.findtext("legis-num", default='N/A'),
        'Vote Question': metadata.findtext("vote-question", default='N/A'),
        'Vote Type': metadata.findtext("vote-type", default='N/A'),
        'Vote Result': metadata.findtext("vote-result", default='N/A'),
        'Action Date': metadata.findtext("action-date", default='N/A'),
        'Action Time': metadata.find(".//action-time").attrib.get("time-etz", 'N/A'),
        'Description': metadata.findtext("vote-desc", default='N/A'),
        'Total Yeas': totals.findtext("yea-total", default='N/A'),
        'Total Nays': totals.findtext("nay-total", default='N/A'),
        'Total Present': totals.findtext("present-total", default='N/A'),
        'Total Not Voting': totals.findtext("not-voting-total", default='N/A')
    }
    return data

def parse_member_votes(xml_data):
    """Extract each member's vote from the XML."""
    member_votes = []
    for vote_element in xml_data.findall(".//recorded-vote"):
        legislator_element = vote_element.find('legislator')
        member_vote_data = {
            'Member Name': legislator_element.text,
            'State': legislator_element.get('state'),
            'Party': legislator_element.get('party'),
            'Vote': vote_element.findtext('vote', default='N/A')
        }
        member_votes.append(member_vote_data)
    return member_votes

def write_headers(sheet, headers):
    """Write the headers to the spreadsheet."""
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column).value = header

def update_spreadsheet_with_member_votes(data, member_votes, file_path):
    """Update the spreadsheet with the new data including each member's vote."""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    headers = ['Majority', 'Congress', 'Session', 'Chamber', 'Roll Call Number', 'Legislation Number',
               'Vote Question', 'Vote Type', 'Vote Result', 'Action Date', 'Action Time',
               'Description', 'Total Yeas', 'Total Nays', 'Total Present', 'Total Not Voting',
               'Member Name', 'State', 'Party', 'Vote']
    
    if sheet.max_row == 1 and all(sheet.cell(row=1, column=col).value is None for col in range(1, len(headers) + 1)):
        write_headers(sheet, headers)
    
    for member_vote in member_votes:
        next_row = sheet.max_row + 1
        for idx, header in enumerate(headers, start=1):
            if header in data:
                sheet.cell(row=next_row, column=idx).value = data.get(header, 'N/A')
            elif header in member_vote:
                sheet.cell(row=next_row, column=idx).value = member_vote.get(header, 'N/A')
    
    wb.save(file_path)

def main(start_roll_call, end_roll_call, file_path, year=2024):
    """Main function to fetch, parse, and update spreadsheet for a range of roll call numbers."""
    for roll_call_number in range(start_roll_call, end_roll_call + 1):
        xml_data = fetch_roll_call_data(roll_call_number, year)
        if xml_data is not None:
            data = parse_xml_data(xml_data)
            member_votes = parse_member_votes(xml_data)
            update_spreadsheet_with_member_votes(data, member_votes, file_path)
            print(f"Successfully updated spreadsheet for roll call {roll_call_number}.")
        else:
            print(f"Data for roll call {roll_call_number} could not be processed.")

if __name__ == "__main__":
    file_path = '/Users/spencer/Desktop/Roll Call/96-104.xlsx'  # Update with your actual path
    start_roll_call = 86  # Use the actual starting roll call number you wish to process
    end_roll_call = 104  # Use the actual ending roll call number you wish to process
    main(start_roll_call, end_roll_call, file_path)
